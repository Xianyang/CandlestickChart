import os
import xlrd, xlwt, xlutils
from xlutils.copy import copy
from datetime import datetime, timedelta
from openpyxl import load_workbook
import blpapi
from pytz import timezone


#------------------------Constants-----------------------
NOW = datetime.now()
hktz = timezone('Asia/Hong_Kong')
esttz = timezone('US/Eastern')
gmttz = timezone('Etc/GMT')
# File Path
file1H = "..\\Intraday\\output.xlsx"
file1D = "..\\Daily\\conf\\\emd_input\\OIL LOG CLOSE.xlsx"
OHLCfileName = "Oil & Refinery Product Price Download_"+NOW.strftime("%d.%m.%Y_%H")+".xlsx"
fileOHLC = "M:\\Data\\Intraday\\"+OHLCfileName
dataFile = "M:\\Chester.Luo\\chartData\\charts1hour.xlsx"

# Which matsuba to use, 3: 1-hour res3, 2: 1-day res2
whichone_1H = 3
whichone_1D = 2

# host and port for connecting to bloomberg
HOST = "localhost"
PORT = 8194
#--------------------end of Constants---------------------


def readChartdate(fileName):
    writeLog("Read last date in chart data")
    book = xlrd.open_workbook(fileName)
    sheet = book.sheet_by_index(0)
    lastr = sheet.nrows-1
    date = sheet.cell(lastr,0).value
    datevalue = datetime(*xlrd.xldate_as_tuple(date, 0))
    datevalue = esttz.localize(datevalue)
    return datevalue


def readOHLC(fileName, chartLastdate):
    DATE, OPEN, HIGH, LOW, CLOSE = 0,0.0,0.0,0.0,0.0
    try:
        book = xlrd.open_workbook(fileName)
    except IOError:
        writeLog("Intraday OHLC not updated")
        return False, DATE, OPEN, HIGH, LOW, CLOSE
    writeLog("Read last hour OHLC")
    sheet = book.sheet_by_index(0)
    lastr = sheet.nrows-1
    for i in range(0,lastr-2):
        row = lastr - i
        date = sheet.cell(row,0).value
        datevalue = datetime(*xlrd.xldate_as_tuple(date, 0))
        localdate = hktz.localize(datevalue)
        DATE = localdate.astimezone(esttz)
        if DATE==chartLastdate:
            break
        if row == 3:
            writeLog("last hour OHLC not available")
            return True, DATE, OPEN, HIGH, LOW, CLOSE
    HIGH = sheet.cell(row,1).value
    LOW = sheet.cell(row,2).value
    CLOSE = sheet.cell(row,3).value
    OPEN = sheet.cell(row,4).value
    return True, DATE, OPEN, HIGH, LOW, CLOSE


def read1H(fileName, whichone):
    writeLog("Read 1-hour extension...")
    book = xlrd.open_workbook(fileName)
    sheet = book.sheet_by_index(0)
    high, low = 0.0,0.0
    date = sheet.cell(1,0).value
    datevalue = datetime(*xlrd.xldate_as_tuple(date, 0)) + timedelta(hours=1)
    localdate = hktz.localize(datevalue)
    DATE = localdate.astimezone(esttz)
    high = sheet.cell(whichone,2).value
    low = sheet.cell(whichone,3).value
    return DATE, high, low


def read1D(fileName, whichone):
    writeLog("Read 1-day extension...")
    book = xlrd.open_workbook(fileName)
    sheet = book.sheet_by_index(0)
    high, low = 0,0
    i = 0
    while sheet.cell_type(i,0) != 0:
        i+=1
    date = sheet.cell(i-1,0).value
    DATE = datetime(*xlrd.xldate_as_tuple(date, 0))
    high = sheet.cell(i,17+whichone).value
    low = sheet.cell(i,20+whichone).value
    return DATE, high, low


def write1D(fileName, date1d, high1d, low1d):
    Wb = xlrd.open_workbook(fileName)
    sheet = Wb.sheet_by_index(0)
    rowNum = sheet.nrows-1
    datevalue = sheet.cell(rowNum,0).value
    date = datetime(*xlrd.xldate_as_tuple(datevalue, 0))
    if date1d.weekday()==4:
        date1d = date1d + timedelta(days=2)
    while date.hour != 19:
        rowNum -= 1
        datevalue = sheet.cell(rowNum,0).value
        date = datetime(*xlrd.xldate_as_tuple(datevalue, 0))
    if date.day == date1d.day:
        writeLog("Write 1-day extension")
        wb = load_workbook(fileName)
        ws = wb.get_sheet_by_name("Sheet1")
        ws.cell(row=rowNum+1,column=8).value = high1d
        ws.cell(row=rowNum+1,column=9).value = low1d
        ws.cell(row=rowNum+1,column=10).value = 24
        wb.save(fileName)



def write1H(fileName, date1h, high1h, low1h):
    if date1h.hour==17:
        date1h = date1h.replace(hour=18)

    endDateTime = NOW
    startDateTime = NOW-timedelta(days=1)
    # Fill SessionOptions
    sessionOptions = blpapi.SessionOptions()
    sessionOptions.setServerHost(HOST)
    sessionOptions.setServerPort(PORT)

    print "Connecting to %s:%s ..." % (HOST, PORT)
    writeLog("Connecting to server host and port")

    # Create a Session
    session = blpapi.Session(sessionOptions)
    
    # Start a Session
    if not session.start():
        print "Failed to start session."
        writeLog("Failed to start session.")
        return
    try:
        # Open service to get historical data from
        if not session.openService("//blp/refdata"):
            print "Failed to open //blp/refdata"
            writeLog("Failed to open //blp/refdata.")
            return
        # Obtain previously opened service
        refDataService = session.getService("//blp/refdata")

        # Create and fill the request for the intraday bar
        request = refDataService.createRequest("IntradayBarRequest")
        request.set("security", "CL1 COMDTY")
        request.set("eventType", "TRADE")
        request.set("startDateTime", startDateTime)
        request.set("endDateTime", endDateTime)
        request.set("interval", 60)

        # send request out
        session.sendRequest(request)
        writeLog("Start session successfully.")
        wb = load_workbook(fileName)
        ws = wb.get_sheet_by_name("Sheet1")
        rowNum = ws.max_row + 1

        while(True):
            ev = session.nextEvent()
            for msg in ev:
                if msg.messageType()=="IntradayBarResponse":
                    barData=msg.getElement("barData")
                    barTickDataArray=barData.getElement("barTickData")
                    numRow=barTickDataArray.numValues()

                    for j in range(0,numRow):
                        barTickData=barTickDataArray.getValueAsElement(j)
                        DATE=barTickData.getElement("time").getValueAsDatetime()
                        openPrice=barTickData.getElement("open").getValueAsFloat()
                        highPrice=barTickData.getElement("high").getValueAsFloat()
                        lowPrice=barTickData.getElement("low").getValueAsFloat()
                        closePrice=barTickData.getElement("close").getValueAsFloat()
                        DATE = gmttz.localize(DATE)
                        d = DATE.astimezone( esttz )
                        if d.hour==date1h.hour:
                            writeLog("Write 1-hour extension and new OHLC")
                            ws.cell(row=rowNum,column=2).value = openPrice
                            ws.cell(row=rowNum,column=3).value = highPrice
                            ws.cell(row=rowNum,column=4).value = lowPrice
                            ws.cell(row=rowNum,column=5).value = closePrice
                            ws.cell(row=rowNum,column=1).value = date1h
                            ws.cell(row=rowNum,column=6).value = high1h
                            ws.cell(row=rowNum,column=7).value = low1h

            if ev.eventType() == blpapi.Event.RESPONSE:
                # Response completly received, so we could exit
                break
    finally:
        #Stop the session
        session.stop()
        print "Waiting for finish..."

    wb.save(fileName)


def writeOHLC(fileName, OPEN, HIGH, LOW, CLOSE):
    writeLog("Write last hour OHLC")
    wb = load_workbook(fileName)
    ws = wb.get_sheet_by_name("Sheet1")
    rowNum = ws.max_row
    ws.cell(row=rowNum,column=2).value = OPEN
    ws.cell(row=rowNum,column=3).value = HIGH
    ws.cell(row=rowNum,column=4).value = LOW
    ws.cell(row=rowNum,column=5).value = CLOSE
    wb.save(fileName)


def writeLog(content):
    logfile = "log1hour.txt"
    f = open(logfile, 'a')
    f.write(str(datetime.now())+' '+str(content)+'\n')
    f.close()


                
if __name__ == "__main__":
    chartLastdate = readChartdate(dataFile)
    writeLog("-----------------------------Chart data update starts---------------------------------")
    flagOHLC, dateOHLC, OPEN, HIGH, LOW, CLOSE = readOHLC(fileOHLC, chartLastdate)
    if flagOHLC==True:
        if dateOHLC == chartLastdate:
            writeOHLC(dataFile, OPEN, HIGH, LOW, CLOSE)
            print "read OHLC: ", dateOHLC, OPEN, HIGH, LOW, CLOSE
        date1h, high1h, low1h = read1H(file1H, whichone_1H)
        print "read 1hour High&Low: ", date1h, high1h, low1h
        if date1h > chartLastdate:
            write1H(dataFile, date1h, high1h, low1h)
        else:
            writeLog("1-hour extension not updated")
    if NOW.hour==7:
        date1d, high1d, low1d = read1D(file1D, whichone_1D)
        print "read 1day High&Low: ", date1d, high1d, low1d
        write1D(dataFile, date1d, high1d, low1d)
