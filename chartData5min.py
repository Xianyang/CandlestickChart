import os
import xlrd, xlwt, xlutils
from xlutils.copy import copy
from datetime import datetime, timedelta
from openpyxl import load_workbook
import blpapi
from pytz import timezone


#------------------------Constants-----------------------
hktz = timezone('Asia/Hong_Kong')
esttz = timezone('US/Eastern')
gmttz = timezone('Etc/GMT')
NOW = datetime.now(hktz).replace(second=0, microsecond=0)
nowEST = NOW.astimezone(esttz)
# File Path
file1H = "..\\Intraday\\output.xlsx"
file1D = "..\\Daily\\conf\\\emd_input\\OIL LOG CLOSE.xlsx"
dataFile = "M:\\Chester.Luo\\chartData\\charts5min.xlsx"

# Which matsuba to use, 3: 1-hour res3, 2: 1-day res2
whichone_1H = 3
whichone_1D = 2

# host and port for connecting to bloomberg
HOST = "localhost"
PORT = 8194
#--------------------end of Constants---------------------


def readChartdate(fileName):
    writeLog("Read last date in chart data")
    book = xlrd.open_workbook(fileName)
    sheet = book.sheet_by_index(0)
    lastr = sheet.nrows-1
    date = sheet.cell(lastr,0).value
    datevalue = datetime(*xlrd.xldate_as_tuple(date,0))
    datevalue = esttz.localize(datevalue)
    return datevalue


# check whether 1-hour matsuba of the hour is recorded in the data file, true-recorded, false-not recorded
def checkHx(fileName, newLastdate):
    writeLog("Check whether 1-hour matsuba of the hour is recorded in the data file...")
    book = xlrd.open_workbook(fileName)
    sheet = book.sheet_by_index(0)
    rowNum = sheet.nrows-1
    datevalue = newLastdate
    while datevalue.hour == newLastdate.hour:
        if sheet.cell(rowNum,8).value != '':
            writeLog("1-hour matsuba of the hour is recorded")
            return True
        rowNum -= 1
        date = sheet.cell(rowNum,0).value
        datevalue = datetime(*xlrd.xldate_as_tuple(date,0))
    writeLog("1-hour matsuba of the hour not recorded")
    return False


def read1H(fileName, whichone):
    writeLog("Read 1-hour extension...")
    book = xlrd.open_workbook(fileName)
    sheet = book.sheet_by_index(0)
    high, low = 0.0,0.0
    date = sheet.cell(1,0).value
    datevalue = datetime(*xlrd.xldate_as_tuple(date,0)) + timedelta(hours=1)
    localdate = hktz.localize(datevalue)
    DATE = localdate.astimezone(esttz)
    high = sheet.cell(whichone,2).value
    low = sheet.cell(whichone,3).value
    return DATE, high, low


def read1D(fileName, whichone):
    writeLog("Read 1-day extension...")
    book = xlrd.open_workbook(fileName)
    sheet = book.sheet_by_index(0)
    high, low = 0,0
    i = 0
    while sheet.cell_type(i,0) != 0:
        i+=1
    date = sheet.cell(i-1,0).value
    DATE = datetime(*xlrd.xldate_as_tuple(date, 0))
    high = sheet.cell(i,17+whichone).value
    low = sheet.cell(i,20+whichone).value
    return DATE, high, low


def write1D(fileName, date1d, high1d, low1d):
    Wb = xlrd.open_workbook(fileName)
    sheet = Wb.sheet_by_index(0)
    rowNum = sheet.nrows-1
    datevalue = sheet.cell(rowNum,0).value
    date = datetime(*xlrd.xldate_as_tuple(datevalue, 0))
    while date.hour != 18 or date.minute != 0:
        rowNum -= 1
        datevalue = sheet.cell(rowNum,0).value
        date = datetime(*xlrd.xldate_as_tuple(datevalue, 0))
    if date >= date1d:
        writeLog("Write 1-day extension")
        wb = load_workbook(fileName)
        ws = wb.get_sheet_by_name("Sheet1")
        ws.cell(row=rowNum+1,column=11).value = high1d
        ws.cell(row=rowNum+1,column=12).value = low1d
        ws.cell(row=rowNum+1,column=13).value = 288
        wb.save(fileName)


def write1H(fileName, high1h, low1h, duration):
    writeLog("Write 1-hour extension")
    wb = load_workbook(fileName)
    ws = wb.get_sheet_by_name("Sheet1")
    rowNum = ws.max_row
    ws.cell(row=rowNum,column=8).value = high1h
    ws.cell(row=rowNum,column=9).value = low1h
    ws.cell(row=rowNum,column=10).value = duration
    wb.save(fileName)


def write5min(fileName, chartLastdate):
    newLastdate = chartLastdate
    endDateTime = NOW
    startDateTime = NOW-timedelta(days=1)
    # Fill SessionOptions
    sessionOptions = blpapi.SessionOptions()
    sessionOptions.setServerHost(HOST)
    sessionOptions.setServerPort(PORT)

    print "Connecting to %s:%s ..." % (HOST, PORT)
    writeLog("Connecting to server host and port")

    # Create a Session
    session = blpapi.Session(sessionOptions)
    
    # Start a Session
    if not session.start():
        print "Failed to start session."
        writeLog("Failed to start session.")
        return chartLastdate
    try:
        # Open service to get historical data from
        if not session.openService("//blp/refdata"):
            print "Failed to open //blp/refdata"
            writeLog("Failed to open //blp/refdata.")
            return chartLastdate
        # Obtain previously opened service
        refDataService = session.getService("//blp/refdata")

        # Create and fill the request for the intraday bar
        request = refDataService.createRequest("IntradayBarRequest")
        request.set("security", "CL1 COMDTY")
        request.set("eventType", "TRADE")
        request.set("startDateTime", startDateTime)
        request.set("endDateTime", endDateTime)
        request.set("interval", 5)

        # send request out
        session.sendRequest(request)
        writeLog("Start session successfully.")
        wb = load_workbook(fileName)
        ws = wb.get_sheet_by_name("Sheet1")
        rowNum = ws.max_row + 1

        while(True):
            ev = session.nextEvent()
            for msg in ev:
                if msg.messageType()=="IntradayBarResponse":
                    barData=msg.getElement("barData")
                    barTickDataArray=barData.getElement("barTickData")
                    numRow=barTickDataArray.numValues()

                    for j in range(0,numRow):
                        barTickData=barTickDataArray.getValueAsElement(j)
                        DATE=barTickData.getElement("time").getValueAsDatetime()
                        openPrice=barTickData.getElement("open").getValueAsFloat()
                        highPrice=barTickData.getElement("high").getValueAsFloat()
                        lowPrice=barTickData.getElement("low").getValueAsFloat()
                        closePrice=barTickData.getElement("close").getValueAsFloat()
                        DATE = gmttz.localize(DATE)
                        d = DATE.astimezone( esttz )
                        if d > newLastdate and d < nowEST:
                            writeLog("Write 5-min OHLC")
                            newLastdate = d
                            ws.cell(row=rowNum,column=1).value = d
                            ws.cell(row=rowNum,column=2).value = openPrice
                            ws.cell(row=rowNum,column=3).value = highPrice
                            ws.cell(row=rowNum,column=4).value = lowPrice
                            ws.cell(row=rowNum,column=5).value = closePrice
                            rowNum += 1
                            
            if ev.eventType() == blpapi.Event.RESPONSE:
                # Response completly received, so we could exit
                break
    finally:
        #Stop the session
        session.stop()
        print "Waiting for finish..."

    wb.save(fileName)
    return newLastdate


def writeLog(content):
    logfile = "log5min.txt"
    f = open(logfile, 'a')
    f.write(str(datetime.now())+' '+str(content)+'\n')
    f.close()


if __name__ == "__main__":
    writeLog("-----------------------------Chart data update starts---------------------------------")
    chartLastdate = readChartdate(dataFile)
    newLastdate = write5min(dataFile, chartLastdate)
    updateFlag = checkHx(dataFile, newLastdate)
    if not updateFlag:
        date1h, high1h, low1h = read1H(file1H, whichone_1H)
        if date1h.hour==17:
            date1h = date1h.replace(hour=18)
        if date1h.hour == newLastdate.hour:
            duration = 12-int(newLastdate.minute/5)
            write1H(dataFile, high1h, low1h, duration)
    if NOW.hour==7:
        date1d, high1d, low1d = read1D(file1D, whichone_1D)
        print "read 1day High&Low: ", date1d, high1d, low1d
        write1D(dataFile, date1d, high1d, low1d)
